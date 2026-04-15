"""
Magellan Microplate Reader 원본 데이터 → 가공 엑셀 변환 모듈
업데이트된 process_growth_curve.py 로직을 웹 서비스에서 사용 가능하도록 모듈화.

주요 변경사항:
- SAMPLE_MAP 지원 (SM 그룹코드 → 샘플명/펩톤 농도 매핑)
- 종합 시트에 샘플 매핑 테이블 추가
- raw 시트에 그룹코드, 샘플명, 펩톤 농도 컬럼 추가
- data 가공 시트에 균주명, 그룹코드, 샘플명, 펩톤 농도 컬럼 추가
"""

import io
import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# 샘플 매핑 헬퍼
# ---------------------------------------------------------------------------
def get_display_name(group_code: str, sample_map: dict) -> str:
    """SM 그룹코드 → 표시용 샘플명 변환."""
    if group_code in sample_map:
        return sample_map[group_code][0]
    return group_code


def get_peptone_pct(group_code: str, sample_map: dict):
    """SM 그룹코드 → 펩톤 농도(%) 반환. 매핑 없으면 None."""
    if group_code in sample_map:
        return sample_map[group_code][1]
    return None


def get_strain(group_code: str, sample_map: dict, fallback: str = "") -> str:
    """SM 그룹코드 → 균주명 반환. 매핑 없으면 fallback 사용."""
    if group_code in sample_map:
        return sample_map[group_code][2] or fallback
    return fallback


# ---------------------------------------------------------------------------
# 1. 원본 데이터 읽기 (첫 번째 Raw OD 블록만 추출)
# ---------------------------------------------------------------------------
def read_raw_block(filepath_or_bytes):
    """첫 번째 시트에서 Raw OD 데이터 블록만 추출.

    Parameters
    ----------
    filepath_or_bytes : str | bytes | io.BytesIO
        파일 경로 또는 파일 바이트

    Returns
    -------
    df : DataFrame with columns: Well, Sample, T0, T1, T2, ...
    time_seconds : list[int] — 각 타임포인트의 초 단위 값
    original_raw : DataFrame — 원본 시트 전체 데이터 (Tecan Raw 보존용)
    """
    if isinstance(filepath_or_bytes, bytes):
        filepath_or_bytes = io.BytesIO(filepath_or_bytes)

    wb = pd.ExcelFile(filepath_or_bytes, engine="openpyxl")
    sheet_name = wb.sheet_names[0]
    raw = pd.read_excel(
        filepath_or_bytes, sheet_name=sheet_name, header=None, engine="openpyxl"
    )

    # ── 원본 데이터 보존 (Transpose 되기 전의 원본 형태) ──
    original_raw_preserve = raw.copy()

    # ── 세로형(수직형) 데이터 감지 및 가로형(수평형)으로 변환 ──
    # 첫 번째 열(Col 0)에 시간 데이터('숫자s')가 3개 이상 존재하면 세로형으로 판단
    col_0_strs = raw.iloc[:, 0].dropna().astype(str)
    if sum(col_0_strs.str.match(r"^\d+\s*s$")) > 2:
        raw = raw.T.reset_index(drop=True)
        # 컬럼명을 0, 1, 2... 순서로 다시 맞춤
        raw.columns = range(raw.shape[1])

    # 첫 번째 행: 시간 헤더 탐색 (숫자 + 's' 또는 '숫자 s' 패턴)
    header_row_idx = None
    for i, row in raw.iterrows():
        vals = row.dropna().astype(str).tolist()
        if any(re.match(r"^\d+\s*s$", v) for v in vals):
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("시간 헤더 행(예: '0s', '3593s' ...)을 찾을 수 없습니다.")

    # 데이터 행에서 실제 OD가 있는 범위 결정
    first_data_row = header_row_idx + 1
    if first_data_row >= len(raw):
        raise ValueError("데이터 행을 찾을 수 없습니다.")

    # 첫 데이터 행의 non-null 개수로 실제 시간 포인트 수 결정
    first_od = raw.iloc[first_data_row, 2:]
    n_times = first_od.notna().sum()

    # 시간(초) 추출 — 처음 n_times개만 사용
    raw_headers = raw.iloc[header_row_idx, 2 : 2 + n_times].tolist()
    time_seconds = []
    for h in raw_headers:
        m = re.match(r"(\d+)\s*s", str(h))
        time_seconds.append(int(m.group(1)) if m else 0)

    # ── 단조 증가 검사: 시간이 다시 작아지면 (두 번째 블록 시작) 거기서 잘라냄 ──
    # Magellan 리더가 여러 측정 블록을 연속으로 내보내는 경우 방지
    cut_idx = len(time_seconds)
    for idx in range(1, len(time_seconds)):
        if time_seconds[idx] <= time_seconds[idx - 1]:
            cut_idx = idx
            break
    if cut_idx < len(time_seconds):
        time_seconds = time_seconds[:cut_idx]
        n_times = cut_idx

    # 고유 컬럼명 생성 (T0, T1, T2, ...)
    time_cols = [f"T{i}" for i in range(n_times)]

    # 데이터 행 추출
    rows = []
    for i in range(first_data_row, len(raw)):
        well = raw.iloc[i, 0]
        sample = raw.iloc[i, 1]
        if pd.isna(well) or str(well).strip() == "":
            break
        od_values = raw.iloc[i, 2 : 2 + n_times].tolist()
        rows.append([str(well).strip(), str(sample).strip()] + od_values)

    if not rows:
        raise ValueError("데이터 행을 찾을 수 없습니다.")

    df = pd.DataFrame(rows, columns=["Well", "Sample"] + time_cols)
    for col in time_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df, time_seconds, original_raw_preserve


# ---------------------------------------------------------------------------
# 2. Blank 보정
# ---------------------------------------------------------------------------
def blank_correct(df: pd.DataFrame) -> pd.DataFrame:
    """BL 샘플 평균을 SM 샘플에서 차감한 보정 DataFrame 반환."""
    time_cols = [c for c in df.columns if c.startswith("T")]

    bl_mask = df["Sample"].str.contains("BL", case=False, na=False)
    sm_mask = df["Sample"].str.contains("SM", case=False, na=False)

    if bl_mask.sum() == 0:
        raise ValueError("Blank(BL) 샘플을 찾을 수 없습니다.")
    if sm_mask.sum() == 0:
        raise ValueError("Sample(SM) 샘플을 찾을 수 없습니다.")

    bl_mean = df.loc[bl_mask, time_cols].mean(axis=0)
    corrected = df.loc[sm_mask].copy()
    corrected[time_cols] = corrected[time_cols].subtract(bl_mean, axis=1)

    return corrected


# ---------------------------------------------------------------------------
# 3. 그룹별 통계 (Mean / SD)
# ---------------------------------------------------------------------------
def extract_group_name(sample: str) -> str:
    """SM1_1 → SM1, SM12_3 → SM12 (마지막 _숫자 접미사 제거)."""
    m = re.match(r"^(.+?)_\d+$", sample)
    return m.group(1) if m else sample


def natural_sort_key(s: str):
    """문자열 내의 숫자를 인식하여 자연스러운 정렬(Natural Sort)을 하기 위한 키 함수."""
    import re
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]


def compute_group_stats(df: pd.DataFrame):
    """그룹별 평균(Mean)과 표준편차(SD)를 반환."""
    time_cols = [c for c in df.columns if c.startswith("T")]
    df = df.copy()
    df["Group"] = df["Sample"].apply(extract_group_name)

    mean_df = df.groupby("Group")[time_cols].mean()
    sd_df = df.groupby("Group")[time_cols].std(ddof=1)  # 표본 표준편차

    # 자연스러운 정렬 적용 (SM1, SM2, ..., SM10)
    sorted_groups = sorted(mean_df.index.tolist(), key=natural_sort_key)
    mean_df = mean_df.reindex(sorted_groups)
    sd_df = sd_df.reindex(sorted_groups)

    # 단일 반복인 경우 SD = 0 처리
    sd_df = sd_df.fillna(0)

    return mean_df, sd_df


# ---------------------------------------------------------------------------
# 4. 엑셀 출력
# ---------------------------------------------------------------------------
def write_output_bytes(
    corrected_df: pd.DataFrame,
    mean_df: pd.DataFrame,
    sd_df: pd.DataFrame,
    time_cols: list,
    time_seconds: list,
    metadata: Optional[Dict[str, str]] = None,
    sample_map: Optional[Dict[str, Tuple[str, float, str]]] = None,
    original_raw: Optional[pd.DataFrame] = None,
) -> bytes:
    """3개 시트를 가진 결과 엑셀 파일을 바이트로 생성하여 반환."""
    wb = Workbook()
    meta = metadata or {}
    smap = sample_map or {}
    strain_name = meta.get("strain", "")

    # ── 스타일 정의 ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    groups = mean_df.index.tolist()

    # =================================================================
    # Sheet 1: 종합 (메타데이터 양식 + 샘플 매핑 테이블)
    # =================================================================
    ws1 = wb.active
    ws1.title = "종합"

    meta_fields = [
        ("일자", meta.get("experiment_date", "")),
        ("배경", ""),
        ("목표", meta.get("goal", "")),
        ("방법", ""),
        ("균주", strain_name),
        ("배지", ""),
        ("배양조건", ""),
        ("장비", "Magellan Microplate Reader"),
        ("비고", ""),
    ]
    for i, (label, default) in enumerate(meta_fields, start=1):
        cell_label = ws1.cell(row=i, column=1, value=label)
        cell_label.font = header_font
        cell_label.fill = header_fill
        cell_label.border = thin_border
        cell_label.alignment = center_align
        cell_val = ws1.cell(row=i, column=2, value=default)
        cell_val.border = thin_border
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 60

    # ── 샘플 매핑 테이블 ──
    map_start_row = len(meta_fields) + 3  # 빈 행 하나 띄움

    map_title = ws1.cell(row=map_start_row, column=1, value="실험 샘플 구성")
    map_title.font = Font(bold=True, size=12)

    map_headers = ["그룹코드", "펩톤1", "비율1(%)", "펩톤2", "비율2(%)", "총농도(%)"]
    map_header_fill = PatternFill("solid", fgColor="B4C6E7")
    blend_fill = PatternFill("solid", fgColor="E2EFDA")
    for j, h in enumerate(map_headers, start=1):
        cell = ws1.cell(row=map_start_row + 1, column=j, value=h)
        cell.font = header_font
        cell.fill = map_header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for i, grp in enumerate(groups, start=map_start_row + 2):
        # 그룹코드
        cell_code = ws1.cell(row=i, column=1, value=grp)
        cell_code.border = thin_border
        cell_code.alignment = center_align

        # 블렌딩 정보 확인
        blend = BLEND_INFO.get(grp)
        if blend:
            # 블렌딩 행
            ws1.cell(row=i, column=2, value=blend["peptone_1"]).border = thin_border
            c_r1 = ws1.cell(row=i, column=3, value=blend["ratio_1"])
            c_r1.border = thin_border
            c_r1.alignment = center_align
            ws1.cell(row=i, column=4, value=blend["peptone_2"]).border = thin_border
            c_r2 = ws1.cell(row=i, column=5, value=blend["ratio_2"])
            c_r2.border = thin_border
            c_r2.alignment = center_align
            # 블렌딩 행 하이라이트
            for col in range(1, 7):
                ws1.cell(row=i, column=col).fill = blend_fill
        else:
            # 단일 펩톤 행
            display = get_display_name(grp, smap)
            ws1.cell(
                row=i, column=2, value=display if display != grp else ""
            ).border = thin_border
            ws1.cell(row=i, column=3, value=100).border = thin_border
            ws1.cell(row=i, column=3).alignment = center_align
            ws1.cell(row=i, column=4, value="").border = thin_border
            ws1.cell(row=i, column=5, value="").border = thin_border

        # 총 펩톤 농도
        pct = get_peptone_pct(grp, smap)
        cell_pct = ws1.cell(row=i, column=6, value=pct if pct is not None else "")
        cell_pct.border = thin_border
        cell_pct.alignment = center_align
        if pct is not None:
            cell_pct.number_format = "0.0"

    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12

    # =================================================================
    # Sheet 2: raw (Blank 보정 Well별 데이터)
    # =================================================================
    ws2 = wb.create_sheet(title="raw")

    # 시간 인덱스 (시간 단위) 계산
    time_hours = [s / 3600 for s in time_seconds]

    # 헤더
    raw_headers = ["Well", "그룹코드", "샘플명", "펩톤 (%)"]
    for j, h in enumerate(raw_headers, start=1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    time_start_col = len(raw_headers) + 1
    for j, t in enumerate(time_cols):
        h = time_hours[j]
        cell = ws2.cell(row=1, column=time_start_col + j, value=f"{h:.2f}h")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 데이터
    for i, (_, row) in enumerate(corrected_df.iterrows(), start=2):
        grp = extract_group_name(row["Sample"])
        display = get_display_name(grp, smap)
        pct = get_peptone_pct(grp, smap)

        ws2.cell(row=i, column=1, value=row["Well"]).border = thin_border
        ws2.cell(row=i, column=2, value=row["Sample"]).border = thin_border

        cell_name = ws2.cell(
            row=i, column=3, value=display if display != grp else ""
        )
        cell_name.border = thin_border

        cell_pct = ws2.cell(
            row=i, column=4, value=pct if pct is not None else ""
        )
        cell_pct.border = thin_border
        cell_pct.alignment = center_align

        for j, t in enumerate(time_cols):
            cell = ws2.cell(
                row=i, column=time_start_col + j, value=round(row[t], 5)
            )
            cell.border = thin_border
            cell.number_format = "0.00000"

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 12

    # =================================================================
    # Sheet 3: data 가공 (Wide-form: Mean | SD)
    # =================================================================
    ws3 = wb.create_sheet(title="data 가공")

    n_times = len(time_cols)

    # 행 1: 구분 라벨
    info_headers = ["균주명", "그룹코드", "샘플명", "펩톤 (%)"]
    for j, h in enumerate(info_headers, start=1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    mean_start_col = len(info_headers) + 1

    # Mean 영역 헤더
    mean_label_cell = ws3.cell(row=1, column=mean_start_col, value="Mean")
    mean_label_cell.font = Font(bold=True, color="FFFFFF")
    mean_label_cell.fill = PatternFill("solid", fgColor="4472C4")
    mean_label_cell.alignment = center_align
    if n_times > 1:
        ws3.merge_cells(
            start_row=1,
            start_column=mean_start_col,
            end_row=1,
            end_column=mean_start_col + n_times - 1,
        )

    # SD 영역 헤더
    sd_start_col = mean_start_col + n_times
    sd_label_cell = ws3.cell(row=1, column=sd_start_col, value="SD")
    sd_label_cell.font = Font(bold=True, color="FFFFFF")
    sd_label_cell.fill = PatternFill("solid", fgColor="ED7D31")
    sd_label_cell.alignment = center_align
    if n_times > 1:
        ws3.merge_cells(
            start_row=1,
            start_column=sd_start_col,
            end_row=1,
            end_column=sd_start_col + n_times - 1,
        )

    # 행 2: 시간 인덱스
    for j, h in enumerate(info_headers, start=1):
        cell = ws3.cell(row=2, column=j, value="")
        cell.border = thin_border
    ws3.cell(row=2, column=len(info_headers), value="Time Index").font = header_font
    ws3.cell(row=2, column=len(info_headers)).fill = header_fill
    ws3.cell(row=2, column=len(info_headers)).border = thin_border

    for j in range(n_times):
        # Mean 시간 인덱스
        cell_m = ws3.cell(row=2, column=mean_start_col + j, value=j)
        cell_m.font = header_font
        cell_m.fill = PatternFill("solid", fgColor="D6E4F0")
        cell_m.alignment = center_align
        cell_m.border = thin_border
        # SD 시간 인덱스
        cell_s = ws3.cell(row=2, column=sd_start_col + j, value=j)
        cell_s.font = header_font
        cell_s.fill = PatternFill("solid", fgColor="FCE4D6")
        cell_s.alignment = center_align
        cell_s.border = thin_border

    # 데이터 행
    for i, grp in enumerate(groups, start=3):
        display = get_display_name(grp, smap)
        pct = get_peptone_pct(grp, smap)

        # 균주명 (그룹별)
        grp_strain = get_strain(grp, smap, strain_name)
        ws3.cell(
            row=i, column=1, value=grp_strain
        ).border = thin_border

        # 그룹코드
        ws3.cell(row=i, column=2, value=grp).border = thin_border

        # 샘플명
        cell_name = ws3.cell(
            row=i, column=3, value=display if display != grp else ""
        )
        cell_name.border = thin_border

        # 펩톤 농도
        cell_pct = ws3.cell(
            row=i, column=4, value=pct if pct is not None else ""
        )
        cell_pct.border = thin_border
        cell_pct.alignment = center_align

        for j, t in enumerate(time_cols):
            # Mean
            cell_m = ws3.cell(
                row=i,
                column=mean_start_col + j,
                value=round(mean_df.loc[grp, t], 5),
            )
            cell_m.border = thin_border
            cell_m.number_format = "0.00000"
            # SD
            cell_s = ws3.cell(
                row=i,
                column=sd_start_col + j,
                value=round(sd_df.loc[grp, t], 5),
            )
            cell_s.border = thin_border
            cell_s.number_format = "0.00000"

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 12

    # =================================================================
    # Sheet 4: raw (세로) — Blank 보정 Well별 데이터 (전치: 시간=행, Well=열)
    # =================================================================
    ws4 = wb.create_sheet(title="raw (세로)")

    # 보정 데이터에서 Well/Sample 정보 추출
    wells = corrected_df["Well"].tolist()
    samples = corrected_df["Sample"].tolist()

    # ── 행 1: 헤더 — "Time (h)" + 각 Well의 Sample 이름
    cell_h = ws4.cell(row=1, column=1, value="Time (h)")
    cell_h.font = header_font
    cell_h.fill = header_fill
    cell_h.border = thin_border
    cell_h.alignment = center_align

    for ci, sample_name in enumerate(samples, start=2):
        cell = ws4.cell(row=1, column=ci, value=sample_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # ── 행 2: 균주명
    cell_s = ws4.cell(row=2, column=1, value="균주명")
    cell_s.font = header_font
    cell_s.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_s.border = thin_border
    cell_s.alignment = center_align

    for ci, sample_name in enumerate(samples, start=2):
        grp = extract_group_name(sample_name)
        grp_strain = get_strain(grp, smap, strain_name)
        cell = ws4.cell(row=2, column=ci, value=grp_strain)
        cell.border = thin_border
        cell.alignment = center_align

    # ── 행 3: 샘플명
    cell_n = ws4.cell(row=3, column=1, value="샘플명")
    cell_n.font = header_font
    cell_n.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_n.border = thin_border
    cell_n.alignment = center_align

    for ci, sample_name in enumerate(samples, start=2):
        grp = extract_group_name(sample_name)
        display = get_display_name(grp, smap)
        cell = ws4.cell(row=3, column=ci, value=display if display != grp else grp)
        cell.border = thin_border
        cell.alignment = center_align

    # ── 데이터 행: 시간별 OD 값
    data_start_row = 4
    for ti, t in enumerate(time_cols):
        row_idx = data_start_row + ti
        h = time_hours[ti]

        cell_time = ws4.cell(row=row_idx, column=1, value=round(h, 4))
        cell_time.border = thin_border
        cell_time.number_format = "0.00"
        cell_time.alignment = center_align

        for ci, (_, row_data) in enumerate(corrected_df.iterrows(), start=2):
            cell = ws4.cell(row=row_idx, column=ci, value=round(row_data[t], 5))
            cell.border = thin_border
            cell.number_format = "0.00000"

    ws4.column_dimensions["A"].width = 12

    # =================================================================
    # Sheet 5: data 가공 (세로) — 그룹 통계 전치 (시간=행, 그룹=열)
    # =================================================================
    ws5 = wb.create_sheet(title="data 가공 (세로)")

    n_groups = len(groups)

    # ── 행 1: 구분 라벨 — Mean 영역 + SD 영역
    cell_h5 = ws5.cell(row=1, column=1, value="Time (h)")
    cell_h5.font = header_font
    cell_h5.fill = header_fill
    cell_h5.border = thin_border
    cell_h5.alignment = center_align

    mean_col_start = 2
    sd_col_start = 2 + n_groups

    # Mean 구분 라벨 (merge)
    mean_label = ws5.cell(row=1, column=mean_col_start, value="Mean")
    mean_label.font = Font(bold=True, color="FFFFFF")
    mean_label.fill = PatternFill("solid", fgColor="4472C4")
    mean_label.alignment = center_align
    mean_label.border = thin_border
    if n_groups > 1:
        ws5.merge_cells(
            start_row=1,
            start_column=mean_col_start,
            end_row=1,
            end_column=mean_col_start + n_groups - 1,
        )

    # SD 구분 라벨 (merge)
    sd_label = ws5.cell(row=1, column=sd_col_start, value="SD")
    sd_label.font = Font(bold=True, color="FFFFFF")
    sd_label.fill = PatternFill("solid", fgColor="ED7D31")
    sd_label.alignment = center_align
    sd_label.border = thin_border
    if n_groups > 1:
        ws5.merge_cells(
            start_row=1,
            start_column=sd_col_start,
            end_row=1,
            end_column=sd_col_start + n_groups - 1,
        )

    # ── 행 2: 그룹코드
    cell_gc = ws5.cell(row=2, column=1, value="그룹코드")
    cell_gc.font = header_font
    cell_gc.fill = header_fill
    cell_gc.border = thin_border
    cell_gc.alignment = center_align

    for gi, grp in enumerate(groups):
        # Mean 열
        cell_m = ws5.cell(row=2, column=mean_col_start + gi, value=grp)
        cell_m.font = header_font
        cell_m.fill = PatternFill("solid", fgColor="D6E4F0")
        cell_m.border = thin_border
        cell_m.alignment = center_align
        # SD 열
        cell_s = ws5.cell(row=2, column=sd_col_start + gi, value=grp)
        cell_s.font = header_font
        cell_s.fill = PatternFill("solid", fgColor="FCE4D6")
        cell_s.border = thin_border
        cell_s.alignment = center_align

    # ── 행 3: 균주명
    cell_strain = ws5.cell(row=3, column=1, value="균주명")
    cell_strain.font = header_font
    cell_strain.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_strain.border = thin_border
    cell_strain.alignment = center_align

    for gi, grp in enumerate(groups):
        grp_strain = get_strain(grp, smap, strain_name)
        cell_m = ws5.cell(
            row=3, column=mean_col_start + gi,
            value=grp_strain,
        )
        cell_m.border = thin_border
        cell_m.alignment = center_align
        cell_s = ws5.cell(
            row=3, column=sd_col_start + gi,
            value=grp_strain,
        )
        cell_s.border = thin_border
        cell_s.alignment = center_align

    # ── 행 4: 샘플명
    cell_sn = ws5.cell(row=4, column=1, value="샘플명")
    cell_sn.font = header_font
    cell_sn.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_sn.border = thin_border
    cell_sn.alignment = center_align

    for gi, grp in enumerate(groups):
        display = get_display_name(grp, smap)
        disp_val = display if display != grp else grp

        cell_m = ws5.cell(row=4, column=mean_col_start + gi, value=disp_val)
        cell_m.border = thin_border
        cell_m.alignment = center_align
        cell_s = ws5.cell(row=4, column=sd_col_start + gi, value=disp_val)
        cell_s.border = thin_border
        cell_s.alignment = center_align

    # ── 데이터 행: 시간별 Mean/SD 값
    v_data_start = 5
    for ti, t in enumerate(time_cols):
        row_idx = v_data_start + ti
        h = time_hours[ti]

        cell_time = ws5.cell(row=row_idx, column=1, value=round(h, 4))
        cell_time.border = thin_border
        cell_time.number_format = "0.00"
        cell_time.alignment = center_align

        for gi, grp in enumerate(groups):
            # Mean
            cell_m = ws5.cell(
                row=row_idx,
                column=mean_col_start + gi,
                value=round(mean_df.loc[grp, t], 5),
            )
            cell_m.border = thin_border
            cell_m.number_format = "0.00000"
            # SD
            cell_s = ws5.cell(
                row=row_idx,
                column=sd_col_start + gi,
                value=round(sd_df.loc[grp, t], 5),
            )
            cell_s.border = thin_border
            cell_s.number_format = "0.00000"

    ws5.column_dimensions["A"].width = 12

    # =================================================================
    # Sheet 6: 원본 (Tecan Raw) — 업로드된 원본 파일 데이터 전체 보존
    # =================================================================
    if original_raw is not None and not original_raw.empty:
        ws6 = wb.create_sheet(title="원본 (Tecan Raw)")

        raw_fill_header = PatternFill("solid", fgColor="FFF2CC")

        for ri in range(len(original_raw)):
            for ci in range(len(original_raw.columns)):
                val = original_raw.iloc[ri, ci]
                # NaN → 빈 셀
                if pd.isna(val):
                    val = None
                cell = ws6.cell(row=ri + 1, column=ci + 1, value=val)
                cell.border = thin_border
                # 첫 번째 행(시간 헤더)은 스타일 적용
                if ri == 0 and val is not None:
                    cell.font = header_font
                    cell.fill = raw_fill_header
                    cell.alignment = center_align
                # 첫 두 열(Well, Sample)은 스타일 적용
                if ci < 2 and val is not None and ri > 0:
                    cell.font = Font(bold=True)
                    cell.alignment = center_align

    # 바이트로 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ---------------------------------------------------------------------------
# 5. 차트 데이터 추출 (JSON 용)
# ---------------------------------------------------------------------------
def extract_chart_data(
    mean_df: pd.DataFrame,
    sd_df: pd.DataFrame,
    time_seconds: list,
    sample_map: Optional[Dict[str, Tuple[str, float, str]]] = None,
) -> Dict:
    """Plotly 차트에 필요한 데이터를 딕셔너리로 반환."""
    smap = sample_map or {}
    time_cols = [c for c in mean_df.columns if c.startswith("T")]
    time_hours = [round(s / 3600, 2) for s in time_seconds]

    groups = mean_df.index.tolist()
    series = []
    for grp in groups:
        display = get_display_name(grp, smap)
        label = f"{display} ({grp})" if display != grp else grp

        mean_vals = mean_df.loc[grp, time_cols].tolist()
        sd_vals = sd_df.loc[grp, time_cols].tolist()
        # NaN → None for JSON serialization
        mean_vals = [None if (isinstance(v, float) and np.isnan(v)) else round(v, 5) for v in mean_vals]
        sd_vals = [None if (isinstance(v, float) and np.isnan(v)) else round(v, 5) for v in sd_vals]
        series.append({
            "name": label,
            "group_code": grp,
            "mean": mean_vals,
            "sd": sd_vals,
        })

    return {
        "time_hours": time_hours,
        "series": series,
    }


# ---------------------------------------------------------------------------
# 6. 샘플 매핑 파싱 (프론트엔드 JSON → dict)
# ---------------------------------------------------------------------------
def parse_sample_map(sample_map_list: Optional[List[Dict]] = None) -> Dict[str, Tuple[str, float, str]]:
    """프론트엔드에서 전달된 샘플 매핑 리스트를 dict로 변환.

    Parameters
    ----------
    sample_map_list : list of dict
        [{"code": "SM1", "name": "PEA-1", "peptone_pct": 1.0, "strain": "L. plantarum",
          "peptone_1": "PEA-1", "ratio_1": 60, "peptone_2": "SOY-1", "ratio_2": 40}, ...]

    Returns
    -------
    dict: {"SM1": ("PEA-1", 1.0, "L. plantarum"), ...}
    Also stores blend info in BLEND_INFO module-level dict for Excel output.
    """
    global BLEND_INFO
    BLEND_INFO = {}

    if not sample_map_list:
        return {}

    result = {}
    for item in sample_map_list:
        code = item.get("code", "").strip()
        name = item.get("name", "").strip()
        pct = item.get("peptone_pct")
        strain = item.get("strain", "").strip()
        if code and (name or strain):
            try:
                pct_val = float(pct) if pct is not None and pct != "" else 0.0
            except (ValueError, TypeError):
                pct_val = 0.0
            result[code] = (name, pct_val, strain)

            # 블렌딩 정보 저장
            p1 = item.get("peptone_1", "").strip()
            p2 = item.get("peptone_2", "").strip()
            if p1 and p2:
                r1 = float(item.get("ratio_1", 100))
                r2 = float(item.get("ratio_2", 0))
                BLEND_INFO[code] = {
                    "peptone_1": p1, "ratio_1": r1,
                    "peptone_2": p2, "ratio_2": r2,
                }

    return result


# 블렌딩 정보 저장용 모듈 변수
BLEND_INFO: Dict[str, Dict] = {}


# ---------------------------------------------------------------------------
# 통합 처리 함수
# ---------------------------------------------------------------------------
def process_file(
    file_bytes: bytes,
    metadata: Optional[Dict[str, str]] = None,
    sample_map_list: Optional[List[Dict]] = None,
) -> Tuple[bytes, Dict]:
    """
    원본 엑셀 바이트 → (가공 Excel 바이트, 차트 JSON 데이터) 반환.

    Parameters
    ----------
    file_bytes : bytes
        업로드된 원본 엑셀 파일 바이트
    metadata : dict, optional
        {"experiment_date": "...", "goal": "...", "strain": "..."}
    sample_map_list : list of dict, optional
        [{"code": "SM1", "name": "MRS (Control)", "peptone_pct": 0.0}, ...]

    Returns
    -------
    (excel_bytes, chart_data)
    """
    sample_map = parse_sample_map(sample_map_list)

    # 1) Raw 블록 추출
    df, time_seconds, original_raw = read_raw_block(file_bytes)
    time_cols = [c for c in df.columns if c.startswith("T")]

    # 2) Blank 보정
    corrected = blank_correct(df)

    # 3) 그룹 통계
    mean_df, sd_df = compute_group_stats(corrected)

    # 4) 엑셀 출력 (바이트)
    excel_bytes = write_output_bytes(
        corrected, mean_df, sd_df, time_cols, time_seconds, metadata, sample_map,
        original_raw,
    )

    # 5) 차트 데이터
    chart_data = extract_chart_data(mean_df, sd_df, time_seconds, sample_map)

    # 6) 그룹 목록 (프론트엔드 샘플 매핑 UI 용)
    chart_data["groups"] = mean_df.index.tolist()

    return excel_bytes, chart_data


# ---------------------------------------------------------------------------
# 배지 성분 최적화용 처리 함수
# ---------------------------------------------------------------------------
def _variations_to_sample_map(variations: List[Dict]) -> Dict[str, Tuple[str, float, str]]:
    """variations list → sample_map 호환 dict.

    Display name priority: condition_name > description > fallback.
    샘플명 컬럼에는 연구원이 지정한 '실험군 이름'(condition_name)이 표시됩니다.
    """
    global BLEND_INFO
    BLEND_INFO = {}
    smap: Dict[str, Tuple[str, float, str]] = {}
    for var in variations or []:
        code = (var.get("code") or "").strip()
        if not code:
            continue
        strain = (var.get("strain") or "").strip()
        condition = (var.get("condition_name") or "").strip()
        desc = (var.get("description") or "").strip()
        overrides = var.get("overrides") or {}
        # Prefer explicit condition_name; fall back to description; else 'Base'/'Variation'.
        display = condition or desc or ("Base" if not overrides else "Variation")
        smap[code] = (display, 0.0, strain)
    return smap


# ---------------------------------------------------------------------------
# 배지 최적화 전용 엑셀 출력 (종합 시트 4-섹션)
# ---------------------------------------------------------------------------
def _unique_conditions(variations: List[Dict]) -> List[Dict]:
    """condition_name 기준 unique 실험군 리스트 반환 (입력 순서 보존).

    각 entry: {"name": condition_name, "description": ..., "composition": [...],
               "strain": ..., "codes": ["SM2","SM3",...], "overrides": {...} (legacy)}
    """
    seen = {}
    for var in variations or []:
        cond = (var.get("condition_name") or "").strip()
        if not cond:
            cond = (var.get("description") or "").strip() or var.get("code", "")
        if cond not in seen:
            seen[cond] = {
                "name": cond,
                "description": (var.get("description") or "").strip(),
                "composition": var.get("composition") or [],
                "overrides": var.get("overrides") or {},   # legacy fallback
                "strain": (var.get("strain") or "").strip(),
                "codes": [],
            }
        seen[cond]["codes"].append(var.get("code", ""))
    return list(seen.values())


def _composition_to_dict(comp_list: List[Dict]) -> Dict[str, Dict]:
    """[{name, value, unit, category}, ...] → {name: {value, unit, category}}."""
    return {
        c.get("name", ""): {
            "value": c.get("value", 0),
            "unit": c.get("unit", "g/L"),
            "category": c.get("category", "other"),
        }
        for c in (comp_list or [])
        if c.get("name")
    }


def _diff_composition_vs_base(comp_list: List[Dict], base_dict: Dict[str, Dict]) -> List[Dict]:
    """composition vs base 차이 계산.
    Returns list of diff entries: [{name, kind, base_val, base_unit, new_val, new_unit, category}]
    where kind in {'added', 'removed', 'modified'}.
    """
    diffs = []
    comp_dict = _composition_to_dict(comp_list)

    # added or modified
    for name, info in comp_dict.items():
        new_val = info["value"]
        new_unit = info["unit"]
        if name not in base_dict:
            diffs.append({
                "name": name,
                "kind": "added",
                "base_val": None, "base_unit": None,
                "new_val": new_val, "new_unit": new_unit,
                "category": info["category"],
            })
        else:
            b = base_dict[name]
            try:
                same = float(b["value"]) == float(new_val) and b["unit"] == new_unit
            except (TypeError, ValueError):
                same = (b["value"] == new_val) and (b["unit"] == new_unit)
            if not same:
                kind = "removed" if (float(new_val or 0) == 0) else "modified"
                diffs.append({
                    "name": name,
                    "kind": kind,
                    "base_val": b["value"], "base_unit": b["unit"],
                    "new_val": new_val, "new_unit": new_unit,
                    "category": info["category"] or b["category"],
                })

    # removed (in base but missing from composition)
    for name, b in base_dict.items():
        if name not in comp_dict:
            diffs.append({
                "name": name,
                "kind": "removed",
                "base_val": b["value"], "base_unit": b["unit"],
                "new_val": 0, "new_unit": b["unit"],
                "category": b["category"],
            })
    return diffs


def _diff_summary_text(diffs: List[Dict]) -> str:
    """diff list → "MgSO₄: 0.1→0.2, Glucose: 20→0, Vitamin B (added)" 형태 문자열."""
    if not diffs:
        return "(Base와 동일)"
    parts = []
    for d in diffs:
        name = d["name"]
        kind = d["kind"]
        if kind == "added":
            parts.append(f"{name} 추가({d['new_val']} {d['new_unit']})")
        elif kind == "removed":
            parts.append(f"{name} 제거")
        else:  # modified
            parts.append(f"{name}: {d['base_val']}→{d['new_val']} {d['new_unit']}")
    return ", ".join(parts)


def write_media_optimization_output_bytes(
    corrected_df: pd.DataFrame,
    mean_df: pd.DataFrame,
    sd_df: pd.DataFrame,
    time_cols: list,
    time_seconds: list,
    metadata: Dict[str, str],
    sample_map: Dict[str, Tuple[str, float, str]],
    base_medium: Dict,
    variations: List[Dict],
    original_raw: Optional[pd.DataFrame] = None,
) -> bytes:
    """배지 성분 최적화 실험용 결과 엑셀 생성.

    종합 시트 구조:
    (a) 실험 메타데이터
    (b) Base Medium 조성표
    (c) 실험군 매핑 테이블 (그룹코드 → 실험군 이름 → Override)
    (d) 실험군별 최종 조성 상세표 (unique 기준)

    raw / data 가공 / raw (세로) / data 가공 (세로) / 원본(Tecan Raw) 시트는
    기존 write_output_bytes 와 동일한 레이아웃을 재사용한다.
    """
    meta = metadata or {}
    smap = sample_map or {}
    base_medium = base_medium or {"preset": "CUSTOM", "composition": []}
    variations = variations or []
    strain_name = meta.get("strain", "")

    preset = base_medium.get("preset", "NONE")
    custom_name = (base_medium.get("custom_name") or "").strip()
    composition = base_medium.get("composition", []) or []
    base_defined = bool(composition)   # (b) section 표시 여부

    # Display label for Base Medium heading
    if preset == "NONE" or (not base_defined and preset != "CUSTOM"):
        base_label = "없음 (각 조성 그룹 독립 정의)"
    elif preset == "CUSTOM":
        base_label = f"사용자 정의 배지 — {custom_name}" if custom_name else "사용자 정의 배지"
    else:
        base_label = preset

    wb = Workbook()

    # ── 스타일 ──
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=12, color="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="BDD7EE")
    override_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    groups = mean_df.index.tolist()
    time_hours = [s / 3600 for s in time_seconds]

    # =================================================================
    # Sheet 1: 종합 (4 섹션)
    # =================================================================
    ws1 = wb.active
    ws1.title = "종합"

    row = 1

    # ── (a) 실험 메타데이터 ──
    sec_a = ws1.cell(row=row, column=1, value="(a) 실험 메타데이터")
    sec_a.font = section_font
    sec_a.fill = section_fill
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    meta_fields = [
        ("실험 일자", meta.get("experiment_date", "")),
        ("실험 목표", meta.get("goal", "")),
        ("사용 균주", strain_name),
        ("사용 배지", meta.get("base_media", "")),
        ("실험 유형", "배지 성분 최적화"),
        ("Base Medium", base_label),
        ("장비", "Magellan Microplate Reader"),
    ]
    for label, value in meta_fields:
        c_label = ws1.cell(row=row, column=1, value=label)
        c_label.font = header_font
        c_label.fill = header_fill
        c_label.border = thin_border
        c_label.alignment = center_align
        c_val = ws1.cell(row=row, column=2, value=value)
        c_val.border = thin_border
        c_val.alignment = left_align
        ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    row += 1  # spacer

    category_label = {
        "carbon": "탄소원",
        "nitrogen": "질소원",
        "mineral": "무기염류",
        "other": "기타",
    }
    cat_order = {"carbon": 0, "nitrogen": 1, "mineral": 2, "other": 3}

    # ── (b) Base Medium 조성표 (Base 정의 시에만 표시) ──
    if base_defined:
        sec_b_text = f"(b) Base Medium 조성표 — {base_label}"
        sec_b = ws1.cell(row=row, column=1, value=sec_b_text)
        sec_b.font = section_font
        sec_b.fill = section_fill
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        comp_headers = ["카테고리", "성분명", "값", "단위"]
        for j, h in enumerate(comp_headers, start=1):
            c = ws1.cell(row=row, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = center_align
        row += 1

        comp_sorted = sorted(
            composition,
            key=lambda x: (cat_order.get(x.get("category", "other"), 9), x.get("name", "")),
        )
        for comp in comp_sorted:
            cat = comp.get("category", "other")
            ws1.cell(row=row, column=1, value=category_label.get(cat, cat)).border = thin_border
            ws1.cell(row=row, column=2, value=comp.get("name", "")).border = thin_border
            c_val = ws1.cell(row=row, column=3, value=comp.get("value", 0))
            c_val.border = thin_border
            c_val.alignment = center_align
            c_val.number_format = "0.###"
            c_unit = ws1.cell(row=row, column=4, value=comp.get("unit", ""))
            c_unit.border = thin_border
            c_unit.alignment = center_align
            row += 1

        row += 1  # spacer
    else:
        # Base 없음 안내
        sec_b = ws1.cell(row=row, column=1, value=f"(b) Base Medium — {base_label}")
        sec_b.font = section_font
        sec_b.fill = section_fill
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        note_cell = ws1.cell(
            row=row, column=1,
            value="※ Base Medium 이 정의되지 않았습니다. 각 조성 그룹의 전체 조성이 (d) 섹션에 표시됩니다.",
        )
        note_cell.alignment = left_align
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 2

    # ── (c) 실험군 매핑 테이블 ──
    sec_c = ws1.cell(row=row, column=1, value="(c) 실험군 매핑 테이블")
    sec_c.font = section_font
    sec_c.fill = section_fill
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    last_col_header = "Base 대비 차이" if base_defined else "조성 요약"
    map_headers = ["그룹코드", "실험군 이름", "균주", "변형 설명", last_col_header]
    for j, h in enumerate(map_headers, start=1):
        c = ws1.cell(row=row, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = center_align
    row += 1

    # Build variation_lookup by code
    var_by_code = {v.get("code"): v for v in variations if v.get("code")}
    base_dict = _composition_to_dict(composition)

    for grp in groups:
        var = var_by_code.get(grp, {})
        cond_name = (var.get("condition_name") or "").strip()
        desc = (var.get("description") or "").strip()
        strain_v = (var.get("strain") or "").strip() or strain_name
        var_comp = var.get("composition") or []
        overrides = var.get("overrides") or {}   # legacy fallback

        ws1.cell(row=row, column=1, value=grp).border = thin_border
        ws1.cell(row=row, column=1).alignment = center_align
        ws1.cell(row=row, column=2, value=cond_name or "-").border = thin_border
        ws1.cell(row=row, column=3, value=strain_v).border = thin_border
        ws1.cell(row=row, column=4, value=desc or "-").border = thin_border

        # Summary text: Base 대비 차이 (Base 정의 시) or 전체 성분 수 (Base 없음)
        if base_defined:
            if var_comp:
                diffs = _diff_composition_vs_base(var_comp, base_dict)
                summary_text = _diff_summary_text(diffs)
            elif overrides:
                # legacy overrides format
                parts = [
                    f"{n}: {(o.get('value') if isinstance(o, dict) else o)} "
                    f"{(o.get('unit') if isinstance(o, dict) else 'g/L')}".strip()
                    for n, o in overrides.items()
                ]
                summary_text = ", ".join(parts) if parts else "(Base와 동일)"
            else:
                summary_text = "(Base와 동일)"
        else:
            n_comp = len(var_comp)
            summary_text = f"전체 조성 ({n_comp}개 성분)" if n_comp > 0 else "(조성 미정의)"

        c_ov = ws1.cell(row=row, column=5, value=summary_text)
        c_ov.border = thin_border
        c_ov.alignment = left_align
        row += 1

    row += 1  # spacer

    # ── (d) 실험군별 최종 조성 상세표 ──
    d_title = "(d) 실험군별 최종 조성 상세"
    if base_defined:
        d_title += " (Base 대비 차이는 노란색)"
    sec_d = ws1.cell(row=row, column=1, value=d_title)
    sec_d.font = section_font
    sec_d.fill = section_fill
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    unique_conds = _unique_conditions(variations)
    if not unique_conds:
        ws1.cell(row=row, column=1, value="(조성 그룹이 비어있어 상세를 생성하지 않았습니다.)").alignment = left_align
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
    else:
        for cond in unique_conds:
            cond_name = cond["name"] or "(무명)"
            codes_str = ", ".join(cond["codes"])
            strain_v = cond.get("strain") or "-"
            desc = cond.get("description") or ""
            heading_text = f"▸ {cond_name}   [{codes_str}]   (균주: {strain_v}"
            if desc:
                heading_text += f" / {desc}"
            heading_text += ")"

            sub = ws1.cell(row=row, column=1, value=heading_text)
            sub.font = Font(bold=True, size=11, color="2E75B6")
            ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1

            detail_headers = ["카테고리", "성분명", "값", "단위", "비고"]
            for j, h in enumerate(detail_headers, start=1):
                c = ws1.cell(row=row, column=j, value=h)
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
                c.alignment = center_align
            row += 1

            cond_comp = cond.get("composition") or []
            legacy_overrides = cond.get("overrides") or {}

            # Legacy fallback: old payload with overrides only → synthesize composition
            if not cond_comp and legacy_overrides:
                synth = []
                # start from base
                for c in composition:
                    name = c.get("name", "")
                    if name in legacy_overrides:
                        ov = legacy_overrides[name]
                        ov_val = ov.get("value") if isinstance(ov, dict) else ov
                        ov_unit = (ov.get("unit") if isinstance(ov, dict) else None) or c.get("unit", "g/L")
                        synth.append({
                            "name": name, "value": ov_val, "unit": ov_unit,
                            "category": c.get("category", "other"),
                        })
                    else:
                        synth.append({
                            "name": name, "value": c.get("value", 0),
                            "unit": c.get("unit", "g/L"),
                            "category": c.get("category", "other"),
                        })
                # adds
                for name, ov in legacy_overrides.items():
                    if not any(s["name"] == name for s in synth):
                        ov_val = ov.get("value") if isinstance(ov, dict) else ov
                        ov_unit = (ov.get("unit") if isinstance(ov, dict) else None) or "g/L"
                        synth.append({
                            "name": name, "value": ov_val, "unit": ov_unit,
                            "category": "other",
                        })
                cond_comp = synth

            # Compute diff-vs-base for highlighting (only if base_defined)
            if base_defined and cond_comp:
                diffs = _diff_composition_vs_base(cond_comp, base_dict)
                diff_by_name = {d["name"]: d for d in diffs}
            else:
                diff_by_name = {}

            # Build display list: all components in cond_comp + any removed-from-base entries (diff kind=removed)
            display_names = [c.get("name", "") for c in cond_comp]
            for d_name, d in diff_by_name.items():
                if d["kind"] == "removed" and d_name not in display_names:
                    display_names.append(d_name)

            def sort_key_d(n):
                # Prefer category from comp, fallback to base
                for c in cond_comp:
                    if c.get("name") == n:
                        return (cat_order.get(c.get("category", "other"), 9), n)
                if n in base_dict:
                    return (cat_order.get(base_dict[n]["category"], 9), n)
                return (9, n)

            if not display_names:
                empty_cell = ws1.cell(row=row, column=1, value="(조성 없음)")
                empty_cell.alignment = left_align
                ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                row += 1
            else:
                for name in sorted(display_names, key=sort_key_d):
                    # Find value + unit + category
                    comp_entry = next((c for c in cond_comp if c.get("name") == name), None)
                    diff = diff_by_name.get(name)

                    if comp_entry is not None:
                        val = comp_entry.get("value", 0)
                        unit = comp_entry.get("unit", "g/L")
                        cat = comp_entry.get("category", "other")
                    elif diff and diff["kind"] == "removed":
                        # Removed from base → display 0
                        val = 0
                        unit = diff.get("base_unit") or "g/L"
                        cat = diff.get("category", "other")
                    else:
                        continue

                    note = ""
                    is_diff = False
                    if diff:
                        is_diff = True
                        kind = diff["kind"]
                        if kind == "added":
                            note = "◆ 추가 (Base에 없음)"
                        elif kind == "removed":
                            note = "◆ 제거"
                        elif kind == "modified":
                            note = f"◆ 변경 ({diff['base_val']} → {diff['new_val']})"

                    ws1.cell(row=row, column=1, value=category_label.get(cat, cat)).border = thin_border
                    ws1.cell(row=row, column=2, value=name).border = thin_border
                    c_val = ws1.cell(row=row, column=3, value=val)
                    c_val.border = thin_border
                    c_val.alignment = center_align
                    c_val.number_format = "0.###"
                    c_unit = ws1.cell(row=row, column=4, value=unit)
                    c_unit.border = thin_border
                    c_unit.alignment = center_align
                    c_note = ws1.cell(row=row, column=5, value=note)
                    c_note.border = thin_border
                    c_note.alignment = left_align
                    if is_diff:
                        for col in range(1, 6):
                            ws1.cell(row=row, column=col).fill = override_fill
                    row += 1
            row += 1  # blank spacer between conditions

    # Column widths for 종합 sheet
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 26
    ws1.column_dimensions["E"].width = 40

    # =================================================================
    # Sheet 2~5: raw / data 가공 / raw (세로) / data 가공 (세로)
    # 기존 write_output_bytes 의 시트 2~5 와 동일한 레이아웃을 재사용
    # → 샘플명 컬럼에는 condition_name (smap 의 display) 이 들어감.
    # =================================================================
    # 아래 로직은 write_output_bytes 시트 2~5 를 그대로 복제한 것. (maintenance 단순화)
    # Helper: per-SM summary text (diff vs Base if defined, else n 성분)
    def _per_sm_summary(grp_code: str) -> str:
        var = var_by_code.get(grp_code, {})
        var_comp = var.get("composition") or []
        overrides = var.get("overrides") or {}
        if base_defined:
            if var_comp:
                diffs = _diff_composition_vs_base(var_comp, base_dict)
                return _diff_summary_text(diffs)
            if overrides:
                parts = [
                    f"{n}={(o.get('value') if isinstance(o, dict) else o)}"
                    for n, o in overrides.items()
                ]
                return ", ".join(parts)
            return ""
        else:
            n_comp = len(var_comp)
            return f"{n_comp}개 성분" if n_comp else ""

    summary_col_title = "Base 대비 차이" if base_defined else "조성 요약"

    # ──────── Sheet 2: raw ────────
    ws2 = wb.create_sheet(title="raw")
    raw_headers = ["Well", "그룹코드", "실험군", summary_col_title]
    for j, h in enumerate(raw_headers, start=1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    time_start_col = len(raw_headers) + 1
    for j, t in enumerate(time_cols):
        h = time_hours[j]
        cell = ws2.cell(row=1, column=time_start_col + j, value=f"{h:.2f}h")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for i, (_, r) in enumerate(corrected_df.iterrows(), start=2):
        grp = extract_group_name(r["Sample"])
        display = get_display_name(grp, smap)
        summary_text = _per_sm_summary(grp)

        ws2.cell(row=i, column=1, value=r["Well"]).border = thin_border
        ws2.cell(row=i, column=2, value=r["Sample"]).border = thin_border
        c_disp = ws2.cell(row=i, column=3, value=display if display != grp else grp)
        c_disp.border = thin_border
        c_ov = ws2.cell(row=i, column=4, value=summary_text)
        c_ov.border = thin_border
        c_ov.alignment = left_align

        for j, t in enumerate(time_cols):
            c = ws2.cell(row=i, column=time_start_col + j, value=round(r[t], 5))
            c.border = thin_border
            c.number_format = "0.00000"

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 36

    # ──────── Sheet 3: data 가공 ────────
    ws3 = wb.create_sheet(title="data 가공")
    n_times = len(time_cols)
    info_headers = ["균주명", "그룹코드", "실험군", summary_col_title]
    for j, h in enumerate(info_headers, start=1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = center_align

    mean_start_col = len(info_headers) + 1
    mean_label = ws3.cell(row=1, column=mean_start_col, value="Mean")
    mean_label.font = Font(bold=True, color="FFFFFF")
    mean_label.fill = PatternFill("solid", fgColor="4472C4")
    mean_label.alignment = center_align
    if n_times > 1:
        ws3.merge_cells(
            start_row=1, start_column=mean_start_col,
            end_row=1, end_column=mean_start_col + n_times - 1,
        )

    sd_start_col = mean_start_col + n_times
    sd_label = ws3.cell(row=1, column=sd_start_col, value="SD")
    sd_label.font = Font(bold=True, color="FFFFFF")
    sd_label.fill = PatternFill("solid", fgColor="ED7D31")
    sd_label.alignment = center_align
    if n_times > 1:
        ws3.merge_cells(
            start_row=1, start_column=sd_start_col,
            end_row=1, end_column=sd_start_col + n_times - 1,
        )

    for j in range(n_times):
        cell_m = ws3.cell(row=2, column=mean_start_col + j, value=j)
        cell_m.font = header_font
        cell_m.fill = PatternFill("solid", fgColor="D6E4F0")
        cell_m.alignment = center_align
        cell_m.border = thin_border
        cell_s = ws3.cell(row=2, column=sd_start_col + j, value=j)
        cell_s.font = header_font
        cell_s.fill = PatternFill("solid", fgColor="FCE4D6")
        cell_s.alignment = center_align
        cell_s.border = thin_border

    for i, grp in enumerate(groups, start=3):
        display = get_display_name(grp, smap)
        grp_strain = get_strain(grp, smap, strain_name)
        summary_text = _per_sm_summary(grp)

        ws3.cell(row=i, column=1, value=grp_strain).border = thin_border
        ws3.cell(row=i, column=2, value=grp).border = thin_border
        c_disp = ws3.cell(row=i, column=3, value=display if display != grp else grp)
        c_disp.border = thin_border
        c_ov = ws3.cell(row=i, column=4, value=summary_text)
        c_ov.border = thin_border
        c_ov.alignment = left_align

        for j, t in enumerate(time_cols):
            c_m = ws3.cell(row=i, column=mean_start_col + j, value=round(mean_df.loc[grp, t], 5))
            c_m.border = thin_border
            c_m.number_format = "0.00000"
            c_s = ws3.cell(row=i, column=sd_start_col + j, value=round(sd_df.loc[grp, t], 5))
            c_s.border = thin_border
            c_s.number_format = "0.00000"

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 32

    # ──────── Sheet 4: raw (세로) ────────
    ws4 = wb.create_sheet(title="raw (세로)")
    wells = corrected_df["Well"].tolist()
    samples = corrected_df["Sample"].tolist()

    cell_h = ws4.cell(row=1, column=1, value="Time (h)")
    cell_h.font = header_font
    cell_h.fill = header_fill
    cell_h.border = thin_border
    cell_h.alignment = center_align

    for ci, sample_name in enumerate(samples, start=2):
        cell = ws4.cell(row=1, column=ci, value=sample_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # 균주명 row
    cell_s = ws4.cell(row=2, column=1, value="균주명")
    cell_s.font = header_font
    cell_s.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_s.border = thin_border
    cell_s.alignment = center_align
    for ci, sample_name in enumerate(samples, start=2):
        grp = extract_group_name(sample_name)
        grp_strain = get_strain(grp, smap, strain_name)
        cell = ws4.cell(row=2, column=ci, value=grp_strain)
        cell.border = thin_border
        cell.alignment = center_align

    # 실험군 row
    cell_n = ws4.cell(row=3, column=1, value="실험군")
    cell_n.font = header_font
    cell_n.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_n.border = thin_border
    cell_n.alignment = center_align
    for ci, sample_name in enumerate(samples, start=2):
        grp = extract_group_name(sample_name)
        display = get_display_name(grp, smap)
        cell = ws4.cell(row=3, column=ci, value=display if display != grp else grp)
        cell.border = thin_border
        cell.alignment = center_align

    data_start_row = 4
    for ti, t in enumerate(time_cols):
        row_idx = data_start_row + ti
        h = time_hours[ti]
        cell_time = ws4.cell(row=row_idx, column=1, value=round(h, 4))
        cell_time.border = thin_border
        cell_time.number_format = "0.00"
        cell_time.alignment = center_align
        for ci, (_, row_data) in enumerate(corrected_df.iterrows(), start=2):
            cell = ws4.cell(row=row_idx, column=ci, value=round(row_data[t], 5))
            cell.border = thin_border
            cell.number_format = "0.00000"

    ws4.column_dimensions["A"].width = 12

    # ──────── Sheet 5: data 가공 (세로) ────────
    ws5 = wb.create_sheet(title="data 가공 (세로)")
    n_groups = len(groups)
    cell_h5 = ws5.cell(row=1, column=1, value="Time (h)")
    cell_h5.font = header_font
    cell_h5.fill = header_fill
    cell_h5.border = thin_border
    cell_h5.alignment = center_align

    mean_col_start = 2
    sd_col_start = 2 + n_groups

    mean_label5 = ws5.cell(row=1, column=mean_col_start, value="Mean")
    mean_label5.font = Font(bold=True, color="FFFFFF")
    mean_label5.fill = PatternFill("solid", fgColor="4472C4")
    mean_label5.alignment = center_align
    mean_label5.border = thin_border
    if n_groups > 1:
        ws5.merge_cells(
            start_row=1, start_column=mean_col_start,
            end_row=1, end_column=mean_col_start + n_groups - 1,
        )

    sd_label5 = ws5.cell(row=1, column=sd_col_start, value="SD")
    sd_label5.font = Font(bold=True, color="FFFFFF")
    sd_label5.fill = PatternFill("solid", fgColor="ED7D31")
    sd_label5.alignment = center_align
    sd_label5.border = thin_border
    if n_groups > 1:
        ws5.merge_cells(
            start_row=1, start_column=sd_col_start,
            end_row=1, end_column=sd_col_start + n_groups - 1,
        )

    # 그룹코드 row
    cell_gc = ws5.cell(row=2, column=1, value="그룹코드")
    cell_gc.font = header_font
    cell_gc.fill = header_fill
    cell_gc.border = thin_border
    cell_gc.alignment = center_align
    for gi, grp in enumerate(groups):
        c_m = ws5.cell(row=2, column=mean_col_start + gi, value=grp)
        c_m.font = header_font
        c_m.fill = PatternFill("solid", fgColor="D6E4F0")
        c_m.border = thin_border
        c_m.alignment = center_align
        c_s = ws5.cell(row=2, column=sd_col_start + gi, value=grp)
        c_s.font = header_font
        c_s.fill = PatternFill("solid", fgColor="FCE4D6")
        c_s.border = thin_border
        c_s.alignment = center_align

    # 균주명 row
    cell_st = ws5.cell(row=3, column=1, value="균주명")
    cell_st.font = header_font
    cell_st.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_st.border = thin_border
    cell_st.alignment = center_align
    for gi, grp in enumerate(groups):
        grp_strain = get_strain(grp, smap, strain_name)
        c_m = ws5.cell(row=3, column=mean_col_start + gi, value=grp_strain)
        c_m.border = thin_border
        c_m.alignment = center_align
        c_s = ws5.cell(row=3, column=sd_col_start + gi, value=grp_strain)
        c_s.border = thin_border
        c_s.alignment = center_align

    # 실험군 row
    cell_sn = ws5.cell(row=4, column=1, value="실험군")
    cell_sn.font = header_font
    cell_sn.fill = PatternFill("solid", fgColor="E2EFDA")
    cell_sn.border = thin_border
    cell_sn.alignment = center_align
    for gi, grp in enumerate(groups):
        display = get_display_name(grp, smap)
        disp_val = display if display != grp else grp
        c_m = ws5.cell(row=4, column=mean_col_start + gi, value=disp_val)
        c_m.border = thin_border
        c_m.alignment = center_align
        c_s = ws5.cell(row=4, column=sd_col_start + gi, value=disp_val)
        c_s.border = thin_border
        c_s.alignment = center_align

    v_data_start = 5
    for ti, t in enumerate(time_cols):
        row_idx = v_data_start + ti
        h = time_hours[ti]
        cell_time = ws5.cell(row=row_idx, column=1, value=round(h, 4))
        cell_time.border = thin_border
        cell_time.number_format = "0.00"
        cell_time.alignment = center_align
        for gi, grp in enumerate(groups):
            c_m = ws5.cell(row=row_idx, column=mean_col_start + gi, value=round(mean_df.loc[grp, t], 5))
            c_m.border = thin_border
            c_m.number_format = "0.00000"
            c_s = ws5.cell(row=row_idx, column=sd_col_start + gi, value=round(sd_df.loc[grp, t], 5))
            c_s.border = thin_border
            c_s.number_format = "0.00000"

    ws5.column_dimensions["A"].width = 12

    # ──────── Sheet 6: 원본 (Tecan Raw) ────────
    if original_raw is not None and not original_raw.empty:
        ws6 = wb.create_sheet(title="원본 (Tecan Raw)")
        raw_fill_header = PatternFill("solid", fgColor="FFF2CC")
        for ri in range(len(original_raw)):
            for ci in range(len(original_raw.columns)):
                val = original_raw.iloc[ri, ci]
                if pd.isna(val):
                    val = None
                cell = ws6.cell(row=ri + 1, column=ci + 1, value=val)
                cell.border = thin_border
                if ri == 0 and val is not None:
                    cell.font = header_font
                    cell.fill = raw_fill_header
                    cell.alignment = center_align
                if ci < 2 and val is not None and ri > 0:
                    cell.font = Font(bold=True)
                    cell.alignment = center_align

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def process_media_optimization(
    file_bytes: bytes,
    metadata: Optional[Dict[str, str]] = None,
    base_medium: Optional[Dict] = None,
    variations: Optional[List[Dict]] = None,
    composition_groups: Optional[List[Dict]] = None,
) -> Tuple[bytes, Dict]:
    """배지 성분 최적화 실험용 처리.

    Parameters
    ----------
    file_bytes : bytes
    metadata   : dict, optional
    base_medium : dict, optional
        {"preset": "NONE"|"MRS"|"TSB"|"LB"|"CUSTOM", "custom_name": "...",
         "composition": [{"name","value","unit","category"}, ...]}
    variations : list of dict, optional
        per-SM list (새 포맷):
        [{"code": "SM1", "strain": "LR", "condition_name": "Control",
          "description": "...", "composition": [...]}, ...]
        legacy 포맷:
        [{"code": "SM1", "strain": "LR", "description": "...", "overrides": {}}, ...]
    composition_groups : list of dict, optional
        신 포맷 (UI에서 그룹별로 SM에 매핑):
        [{"id": "cg-1", "name": "Control", "strain": "LR", "description": "...",
          "composition": [...], "applied_samples": ["SM1","SM2","SM3"]}, ...]
        제공되면 variations 를 덮어써서 per-SM 으로 expand.

    Returns
    -------
    (excel_bytes, chart_data)  — chart_data 에 base_medium / variation / composition_groups 메타 포함
    """
    base_medium = base_medium or {"preset": "NONE", "composition": []}
    variations = variations or []
    composition_groups = composition_groups or []

    # composition_groups → per-SM variations expansion
    if composition_groups:
        expanded: List[Dict] = []
        for cg in composition_groups:
            applied = cg.get("applied_samples") or []
            for sm in applied:
                expanded.append({
                    "code": sm,
                    "strain": (cg.get("strain") or "").strip(),
                    "condition_name": (cg.get("name") or "").strip(),
                    "description": (cg.get("description") or "").strip(),
                    "composition": cg.get("composition") or [],
                    "group_id": cg.get("id"),
                })
        if expanded:
            variations = expanded

    # variations → sample_map 형태로 변환 (condition_name 우선 display 로 사용)
    sample_map = _variations_to_sample_map(variations)

    # 1) Raw 추출
    df, time_seconds, original_raw = read_raw_block(file_bytes)
    time_cols = [c for c in df.columns if c.startswith("T")]

    # 2) Blank 보정
    corrected = blank_correct(df)

    # 3) 그룹 통계
    mean_df, sd_df = compute_group_stats(corrected)

    # 4) 엑셀 출력 (배지 최적화 전용 — 종합 4-섹션 + raw/data 시트들)
    excel_bytes = write_media_optimization_output_bytes(
        corrected, mean_df, sd_df, time_cols, time_seconds,
        metadata or {}, sample_map, base_medium, variations,
        original_raw,
    )

    # 5) 차트 데이터 + variation 메타 enrich
    chart_data = extract_chart_data(mean_df, sd_df, time_seconds, sample_map)

    # variations list → dict by code
    var_by_code = {v.get("code"): v for v in variations if v.get("code")}

    for s in chart_data["series"]:
        code = s.get("group_code")
        var = var_by_code.get(code, {})
        s["variation_code"] = code
        s["condition_name"] = (var.get("condition_name") or "").strip()
        s["variation_desc"] = var.get("description", "")
        s["variation_overrides"] = var.get("overrides", {})    # legacy
        s["composition"] = var.get("composition", [])          # full per-SM composition
        s["strain"] = var.get("strain", "")
        s["group_id"] = var.get("group_id")

    chart_data["groups"] = mean_df.index.tolist()
    chart_data["experiment_type"] = "media_optimization"
    chart_data["base_medium"] = base_medium
    chart_data["composition_groups"] = composition_groups

    return excel_bytes, chart_data
