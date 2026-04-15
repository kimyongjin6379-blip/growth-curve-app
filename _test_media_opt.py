"""Standalone verification v2: composition_groups → per-SM variations payload."""
import io
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from processor import process_media_optimization


def make_fake_tecan_raw() -> bytes:
    """Minimal Tecan-style raw sheet: BL + SM1..SM4 × 3 replicates."""
    wb = Workbook()
    ws = wb.active
    time_seconds = [0, 3600, 7200, 10800, 14400, 18000]
    header = ["Well", "Sample"] + [f"{s}s" for s in time_seconds]
    ws.append(header)

    for i in range(1, 4):
        ws.append([f"A{i}", f"BL1_{i}"] + [0.05 + 0.001 * i for _ in time_seconds])

    base_profiles = {
        "SM1": [0.10, 0.20, 0.45, 0.90, 1.30, 1.55],
        "SM2": [0.10, 0.22, 0.50, 1.00, 1.45, 1.70],
        "SM3": [0.10, 0.22, 0.50, 1.00, 1.45, 1.70],
        "SM4": [0.10, 0.12, 0.18, 0.25, 0.30, 0.32],
    }
    row_letter = 'B'
    for sm, profile in base_profiles.items():
        for rep in range(1, 4):
            ws.append(
                [f"{row_letter}{rep}", f"{sm}_{rep}"]
                + [v + 0.005 * rep for v in profile]
            )
        row_letter = chr(ord(row_letter) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    raw_bytes = make_fake_tecan_raw()

    metadata = {
        "experiment_date": "2026-04-15",
        "goal": "Min-1 배지 최적화 검증 (v2 composition_groups)",
        "strain": "LR",
        "base_media": "Min-1",
        "media_type": "media_optimization",
    }

    # Base Medium = CUSTOM "Min-1" (Glucose + YE + MgSO₄ + K₂HPO₄)
    base_medium = {
        "preset": "CUSTOM",
        "custom_name": "Min-1",
        "composition": [
            {"name": "Glucose",        "value": 20,  "unit": "g/L", "category": "carbon"},
            {"name": "Yeast Extract",  "value": 5,   "unit": "g/L", "category": "other"},
            {"name": "MgSO₄",          "value": 0.1, "unit": "g/L", "category": "mineral"},
            {"name": "K₂HPO₄",         "value": 2,   "unit": "g/L", "category": "mineral"},
        ],
    }

    # Composition groups — each a full composition, mapped to SM samples
    composition_groups = [
        {
            "id": "cg-1",
            "name": "Control",
            "strain": "LR",
            "description": "대조군 (Base 동일)",
            "composition": [
                {"name": "Glucose",       "value": 20,  "unit": "g/L", "category": "carbon"},
                {"name": "Yeast Extract", "value": 5,   "unit": "g/L", "category": "other"},
                {"name": "MgSO₄",         "value": 0.1, "unit": "g/L", "category": "mineral"},
                {"name": "K₂HPO₄",        "value": 2,   "unit": "g/L", "category": "mineral"},
            ],
            "applied_samples": ["SM1"],
        },
        {
            "id": "cg-2",
            "name": "+Mg 2x",
            "strain": "LR",
            "description": "Mg 2배 증량 (3반복)",
            "composition": [
                {"name": "Glucose",       "value": 20,  "unit": "g/L", "category": "carbon"},
                {"name": "Yeast Extract", "value": 5,   "unit": "g/L", "category": "other"},
                {"name": "MgSO₄",         "value": 0.2, "unit": "g/L", "category": "mineral"},
                {"name": "K₂HPO₄",        "value": 2,   "unit": "g/L", "category": "mineral"},
            ],
            "applied_samples": ["SM2", "SM3"],
        },
        {
            "id": "cg-3",
            "name": "No Glucose + Vitamin",
            "strain": "LR",
            "description": "탄소원 제거 + Vitamin B1 추가",
            "composition": [
                {"name": "Yeast Extract", "value": 5,    "unit": "g/L",  "category": "other"},
                {"name": "MgSO₄",         "value": 0.1,  "unit": "g/L",  "category": "mineral"},
                {"name": "K₂HPO₄",        "value": 2,    "unit": "g/L",  "category": "mineral"},
                {"name": "Vitamin B1",    "value": 0.01, "unit": "g/L",  "category": "other"},
            ],
            "applied_samples": ["SM4"],
        },
    ]

    excel_bytes, chart_data = process_media_optimization(
        raw_bytes, metadata, base_medium, variations=None,
        composition_groups=composition_groups,
    )

    print(f"✓ Excel generated: {len(excel_bytes)} bytes")
    print(f"✓ chart_data keys: {list(chart_data.keys())}")
    print(f"✓ experiment_type: {chart_data.get('experiment_type')}")
    print(f"✓ base_medium.preset: {chart_data['base_medium']['preset']}")
    print(f"✓ base_medium.custom_name: {chart_data['base_medium'].get('custom_name')}")
    print(f"✓ composition_groups count: {len(chart_data.get('composition_groups', []))}")
    print(f"✓ series count: {len(chart_data['series'])}")
    for s in chart_data['series']:
        comp_n = len(s.get('composition', []))
        print(f"    - {s['group_code']}: name={s['name']!r} condition_name={s.get('condition_name')!r} "
              f"composition=[{comp_n} items] group_id={s.get('group_id')}")

    out_path = Path("_test_media_opt_output.xlsx")
    out_path.write_bytes(excel_bytes)
    wb = load_workbook(out_path)
    print(f"\n✓ Excel sheets: {wb.sheetnames}")

    ws = wb["종합"]
    print(f"\n── 종합 sheet preview (first {min(ws.max_row, 80)} rows) ──")
    for r in range(1, min(ws.max_row + 1, 81)):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        cells_display = [("" if v is None else str(v))[:38] for v in cells]
        print(f"  row {r:2d}: {cells_display}")


if __name__ == "__main__":
    main()
